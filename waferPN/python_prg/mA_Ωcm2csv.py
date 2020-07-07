import pathlib                            #標準ライブラリ
import os                                 #標準ライブラリ
from openpyxl import Workbook             #外部ライブラリ
import pandas as pd                       #外部ライブラリ
import csv                                #標準ライブラリ

lwb = Workbook()            #一覧のワークブックを1wbとして作成
lsh = lwb.active            #デフォルトで作成されるワークシートを選択

path = pathlib.Path("抵抗率_電流グラフ変換/waferPN/data")                           #Pathオブジェクトの生成

list_row = 3
list_col = 1
for pass_obj in path.iterdir():                                             #パスがディレクトリを指定している場合、ディレクトリの中のファイルやフォルダーの名前をPathオブジェクトとして順々に返す。
    if pass_obj.match("*.csv"):                                             #path.iterdir()の戻り値であるパスがcsvファイルであるものなら、以下を実行。
        df = pd.read_csv(pass_obj, header=7, index_col=False, delimiter=r",\s*", engine="python")
        volt = df.iloc[:,1]
        amp = df.iloc[:,14]
        file_name = os.path.splitext(os.path.basename(pass_obj))[0]
        if "SNL" in file_name:
            thick = 0.038
        elif "SNM" in file_name:
            thick = 0.028
        elif "SNH" in file_name:
            thick = 0.04
        elif "SPL" in file_name:
            thick = 0.0525
        elif "SPM" in file_name:
            thick = 0.04
        elif "SPH" in file_name:
            thick = 0.04
        # vdf = volt_data.astype(float)
        # adf = amp_data.astype(float)
        lsh.cell(list_row -2, list_col).value = file_name                                           #csvファイル名
        lsh.cell(list_row - 1, list_col).value = "mA"                                               #mA表記
        for row in range(len(volt)):
            lsh.cell(list_row, list_col).value = amp.values[row] * 1000                             #mA値
            list_row += 1
        list_row = 3
        list_col += 1
        lsh.cell(list_row -2, list_col).value = df.values[0][21]                                    #測定日
        lsh.cell(list_row - 1, list_col).value = "Ω"                                                #Ω表記
        for row in range(len(volt)):
            lsh.cell(list_row, list_col).value = volt.values[row] / amp.values[row]                 #Ω値
            list_row += 1
        list_row = 3
        list_col += 1
        lsh.cell(list_row - 1, list_col).value = "Ωcm"                                              #Ωcm表記
        for row in range(len(volt)):
            lsh.cell(list_row, list_col).value = 4.532 * thick * volt.values[row] / amp.values[row] #Ωcm値
            list_row += 1
        list_row = 3
        list_col += 2

lwb.save("抵抗率_電流グラフ変換/waferPN/data/mA_Ω_ΩcmList.xlsx")